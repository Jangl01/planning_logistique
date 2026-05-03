"""
Planning Logistique — version refonte postes
============================================
Postes fixes   : une personne dédiée, jamais changée.
Postes rotation: pool strict de personnes habilitées → rotation équitable,
                 PERSONNE hors pool ne peut apparaître sur ce poste.
"""

import streamlit as st
import pandas as pd
import plotly.graph_objects as go
from datetime import date, timedelta
import json, os, random, requests
from io import BytesIO
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.cell.cell import MergedCell

# ─────────────────────────────────────────────────────────────────────────────
#  PAGE CONFIG
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(page_title="Planning Logistique", layout="wide",
                   initial_sidebar_state="expanded", page_icon="📦")

# ─────────────────────────────────────────────────────────────────────────────
#  CSS
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Mono:wght@400;500&family=Syne:wght@400;600;700;800&display=swap');

html,body,[class*="css"]    { font-family:'Syne',sans-serif; }
.stApp                      { background:#0f1117; color:#e8e8e8; }
.block-container            { padding-top:1.4rem; }
h1,h2,h3                   { font-family:'Syne',sans-serif; font-weight:800; letter-spacing:-0.03em; }
div[data-testid="stSidebar"]{ background:#0d0f18 !important; border-right:1px solid #1f2937; }

input, textarea,
.stTextInput  > div > div > input,
.stTextArea   > div > div > textarea,
div[data-baseweb="input"] input,
div[data-baseweb="textarea"] textarea {
    background-color:#1e2130 !important; color:#e8e8e8 !important;
    border:1px solid #374151 !important; border-radius:8px !important;
    font-family:'DM Mono',monospace !important; font-size:.82rem !important;
    caret-color:#60a5fa !important;
}
input::placeholder,textarea::placeholder { color:#6b7280 !important; }
input:focus,textarea:focus  {
    border-color:#2563eb !important;
    box-shadow:0 0 0 2px rgba(37,99,235,.25) !important; outline:none !important;
}
div[data-baseweb="select"] > div {
    background-color:#1e2130 !important; border-color:#374151 !important;
    color:#e8e8e8 !important; border-radius:8px !important;
}
div[data-baseweb="select"] span { color:#e8e8e8 !important; }
div[data-baseweb="popover"]     { background:#1e2130 !important; }
li[role="option"]               { background:#1e2130 !important; color:#e8e8e8 !important; }
li[role="option"]:hover         { background:#2563eb !important; }
div[data-baseweb="input"]       { background:#1e2130 !important; }

.stTabs [data-baseweb="tab-list"]{ gap:6px; background:#1a1d27; border-radius:12px; padding:6px; }
.stTabs [data-baseweb="tab"]     { border-radius:8px; padding:8px 18px; font-weight:700;
                                   font-size:.78rem; letter-spacing:.06em; text-transform:uppercase; color:#888; }
.stTabs [aria-selected="true"]   { background:#2563eb !important; color:white !important; }

.stButton > button { background:#2563eb; color:white; border:none; border-radius:8px;
    font-family:'Syne',sans-serif; font-weight:700; font-size:.78rem;
    letter-spacing:.05em; padding:8px 16px; transition:all .15s; }
.stButton > button:hover { background:#1d4ed8; transform:translateY(-1px); }
.stDownloadButton > button { background:#059669 !important; color:white !important;
    border:none !important; border-radius:8px !important;
    font-family:'Syne',sans-serif !important; font-weight:700 !important; }

.stDataEditor,.stDataFrame { border-radius:10px; overflow:hidden; }
div[data-testid="stExpander"] { background:#141720; border:1px solid #1f2937;
    border-radius:10px; margin-bottom:8px; }
div[data-testid="stExpander"] summary { color:#c9d1d9 !important; font-weight:600; }

/* ── utility ── */
.metric-card { background:#1a1d27; border:1px solid #2a2d3e;
               border-radius:12px; padding:16px 20px; text-align:center; }
.metric-card .val { font-size:2rem; font-weight:800; font-family:'DM Mono',monospace; }
.metric-card .lbl { font-size:.7rem; color:#888; text-transform:uppercase;
                    letter-spacing:.1em; margin-top:4px; }
.badge { display:inline-block; padding:2px 10px; border-radius:20px;
         font-size:.7rem; font-weight:700; font-family:'DM Mono',monospace; letter-spacing:.05em; }
.badge-ok    { background:#14532d; color:#4ade80; }
.badge-warn  { background:#713f12; color:#fbbf24; }
.badge-error { background:#7f1d1d; color:#f87171; }
.badge-blue  { background:#0f1f3d; color:#60a5fa; }
.badge-gold  { background:#422006; color:#fcd34d; }
.badge-fix   { background:#1e3a5f; color:#93c5fd; }
.badge-rot   { background:#2d1f5e; color:#c4b5fd; }

.conflict-box { background:#2d1515; border:1px solid #7f1d1d; border-radius:10px;
                padding:12px 16px; margin:6px 0; font-size:.82rem; color:#fca5a5; }
.info-box     { background:#0f1f3d; border:1px solid #1d4ed8; border-radius:10px;
                padding:12px 16px; margin:6px 0; font-size:.82rem; color:#93c5fd; }
.tip-box      { background:#0f2318; border:1px solid #065f46; border-radius:10px;
                padding:12px 16px; margin:6px 0; font-size:.82rem; color:#6ee7b7; }
.sec { font-size:.68rem; font-weight:700; letter-spacing:.14em; text-transform:uppercase;
       color:#4b5563; border-bottom:1px solid #1f2937; padding-bottom:3px; margin-bottom:6px; }

/* ── poste card ── */
.post-card { background:#141720; border:1px solid #1f2937; border-radius:10px;
             padding:12px 14px; margin:6px 0; }
.post-card-fix { border-left:3px solid #3b82f6; }
.post-card-rot { border-left:3px solid #8b5cf6; }
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
#  CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────
SAVE_FILE   = "planning_data.json"
DAYS        = ["Lundi","Mardi","Mercredi","Jeudi","Vendredi"]
STATUTS     = ["Présent","ABS","Congé","Jour Férié"]
POST_COLORS = [
    "#2563eb","#7c3aed","#059669","#d97706","#dc2626",
    "#0891b2","#be185d","#65a30d","#9333ea","#0369a1",
    "#b45309","#15803d","#c026d3","#1d4ed8","#047857",
]

# ─────────────────────────────────────────────────────────────────────────────
#  DATA STRUCTURES
# ─────────────────────────────────────────────────────────────────────────────
# postes: list of dicts
# {
#   "nom":    str,
#   "type":   "fixe" | "rotation",
#   "slots":  int,           # nb personnes à assigner (rotation only, fixe=1)
#   "person": str,           # fixe only
#   "pool":   [str, ...],    # rotation only — STRICT, nobody outside allowed
# }

def default_postes():
    return []

# ─────────────────────────────────────────────────────────────────────────────
#  STATE INIT
# ─────────────────────────────────────────────────────────────────────────────
def _monday(d: date) -> date:
    return d - timedelta(days=d.weekday())

def init_state():
    defaults = {
        "week_offset":      0,
        "base_monday":      _monday(date.today()),
        "team":             [],
        "postes":           [],          # ← NEW unified structure
        "tasks":            [],
        "task_pools":       {},
        "operations":       {},
        "daily_status":     {},
        "no_pair":          [],
        "base_pairs":       [],
        "cumulative_posts": {},
        "cumulative_tasks": {},
        "week_history":     {},
        "df_main":          pd.DataFrame(),
        "df_tasks":         pd.DataFrame(),
        "last_week_key":    "",
        "editor_version":   0,
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v

init_state()

# ─────────────────────────────────────────────────────────────────────────────
#  WEEK HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def current_monday() -> date:
    return st.session_state.base_monday + timedelta(weeks=st.session_state.week_offset)

def week_key(d: date) -> str:
    return d.strftime("%Y-W%V")

# ─────────────────────────────────────────────────────────────────────────────
#  PERSISTENCE
# ─────────────────────────────────────────────────────────────────────────────
def build_save_dict() -> dict:
    return {
        "base_monday":      st.session_state.base_monday.strftime("%Y-%m-%d"),
        "week_offset":      st.session_state.week_offset,
        "team":             st.session_state.team,
        "postes":           st.session_state.postes,
        "tasks":            st.session_state.tasks,
        "task_pools":       st.session_state.task_pools,
        "operations":       st.session_state.operations,
        "daily_status":     st.session_state.daily_status,
        "no_pair":          [list(p) for p in st.session_state.no_pair],
        "base_pairs":       [list(p) for p in st.session_state.base_pairs],
        "cumulative_posts": st.session_state.cumulative_posts,
        "cumulative_tasks": st.session_state.cumulative_tasks,
        "week_history":     st.session_state.week_history,
    }

def commit_planning_to_history():
    """Write current df_main/df_tasks into week_history so save_data() persists them."""
    if not st.session_state.df_main.empty:
        st.session_state.week_history[st.session_state.last_week_key or week_key(current_monday())] = {
            "main":  st.session_state.df_main.to_dict("records"),
            "tasks": st.session_state.df_tasks.to_dict("records") if not st.session_state.df_tasks.empty else [],
        }

def save_data():
    """Commit planning to history, save locally + push to Gist."""
    commit_planning_to_history()
    data = build_save_dict()
    try:
        with open(SAVE_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
    except Exception:
        pass  # on cloud, local write may fail — not critical
    # Push to GitHub Gist if token is configured
    _gist_push(data)

# ─────────────────────────────────────────────────────────────────────────────
#  GITHUB GIST PERSISTENCE
# ─────────────────────────────────────────────────────────────────────────────
GIST_FILENAME = "planning_logistique_data.json"

def _gist_token() -> str:
    """Read token from st.secrets or environment variable."""
    try:
        return st.secrets.get("GIST_TOKEN", "")
    except Exception:
        return os.environ.get("GIST_TOKEN", "")

def _gist_id() -> str:
    try:
        return st.secrets.get("GIST_ID", "")
    except Exception:
        return os.environ.get("GIST_ID", "")

def _gist_headers() -> dict:
    return {
        "Authorization": f"token {_gist_token()}",
        "Accept": "application/vnd.github.v3+json",
    }

def _gist_push(data: dict) -> bool:
    """Push data to GitHub Gist. Returns True on success."""
    token = _gist_token()
    gist_id = _gist_id()
    if not token:
        return False
    payload = {"files": {GIST_FILENAME: {"content": json.dumps(data, indent=2, ensure_ascii=False)}}}
    try:
        if gist_id:
            r = requests.patch(f"https://api.github.com/gists/{gist_id}",
                               headers=_gist_headers(), json=payload, timeout=8)
        else:
            payload["description"] = "Planning Logistique — sauvegarde automatique"
            payload["public"] = False
            r = requests.post("https://api.github.com/gists",
                              headers=_gist_headers(), json=payload, timeout=8)
            if r.status_code == 201:
                new_id = r.json().get("id", "")
                # Store for this session so next saves use PATCH
                os.environ["GIST_ID"] = new_id
        return r.status_code in (200, 201)
    except Exception:
        return False

def _gist_pull() -> dict | None:
    """Pull latest data from GitHub Gist. Returns dict or None."""
    token = _gist_token()
    gist_id = _gist_id()
    if not token or not gist_id:
        return None
    try:
        r = requests.get(f"https://api.github.com/gists/{gist_id}",
                         headers=_gist_headers(), timeout=8)
        if r.status_code == 200:
            files = r.json().get("files", {})
            if GIST_FILENAME in files:
                raw_url = files[GIST_FILENAME].get("raw_url", "")
                if raw_url:
                    r2 = requests.get(raw_url, timeout=8)
                    if r2.status_code == 200:
                        return r2.json()
    except Exception:
        pass
    return None

def migrate_old_format(data: dict):
    """Convert old fixed_posts/rotation_posts/rotation_pools to new postes list."""
    if "postes" in data:
        return data.get("postes", [])
    postes = []
    fixed_posts    = data.get("fixed_posts", {})
    rotation_posts = data.get("rotation_posts", {})
    rotation_pools = data.get("rotation_pools", {})
    for nom, person in fixed_posts.items():
        postes.append({"nom": nom, "type": "fixe", "slots": 1, "person": person, "pool": []})
    for nom, slots in rotation_posts.items():
        pool = rotation_pools.get(nom, [])
        postes.append({"nom": nom, "type": "rotation", "slots": slots, "person": "", "pool": pool})
    return postes

def load_data():
    """Load from Gist first (if configured), then fall back to local file."""
    data = None
    # Try Gist first
    if _gist_token() and _gist_id():
        data = _gist_pull()
    # Fall back to local file
    if data is None:
        if not os.path.exists(SAVE_FILE):
            return
        try:
            with open(SAVE_FILE, encoding="utf-8") as f:
                data = json.load(f)
        except Exception:
            st.warning("Fichier corrompu — réinitialisation.")
            return
    bm = data.get("base_monday", str(_monday(date.today())))
    st.session_state.base_monday      = date.fromisoformat(bm)
    st.session_state.week_offset      = data.get("week_offset", 0)
    st.session_state.team             = data.get("team", [])
    st.session_state.postes           = migrate_old_format(data)
    st.session_state.tasks            = data.get("tasks", [])
    st.session_state.task_pools       = data.get("task_pools", {})
    st.session_state.operations       = data.get("operations", {})
    st.session_state.daily_status     = data.get("daily_status", {})
    st.session_state.no_pair          = [tuple(p) for p in data.get("no_pair", [])]
    st.session_state.base_pairs       = [tuple(p) for p in data.get("base_pairs", [])]
    st.session_state.cumulative_posts = data.get("cumulative_posts", {})
    st.session_state.cumulative_tasks = data.get("cumulative_tasks", {})
    st.session_state.week_history     = data.get("week_history", {})

load_data()

# ─────────────────────────────────────────────────────────────────────────────
#  HELPERS — parsing
# ─────────────────────────────────────────────────────────────────────────────
def parse_pairs(text: str) -> list:
    pairs = []
    for line in text.splitlines():
        parts = [x.strip() for x in line.split(",")]
        if len(parts) == 2 and all(parts):
            pairs.append(tuple(parts))
    return pairs

def parse_list(text: str) -> list:
    return [x.strip() for x in text.replace(",", "\n").splitlines() if x.strip()]

def parse_dict_list(text: str) -> dict:
    result = {}
    for line in text.splitlines():
        line = line.strip()
        if ":" not in line: continue
        k, _, v = line.partition(":")
        k = k.strip()
        if k: result[k] = [x.strip() for x in v.split(",") if x.strip()]
    return result

# ─────────────────────────────────────────────────────────────────────────────
#  HELPERS — counters
# ─────────────────────────────────────────────────────────────────────────────
def get_cum_post(post: str, person: str) -> int:
    return st.session_state.cumulative_posts.get(f"{post}|{person}", 0)

def inc_cum_post(post: str, person: str):
    k = f"{post}|{person}"
    st.session_state.cumulative_posts[k] = st.session_state.cumulative_posts.get(k, 0) + 1

def get_cum_task(task: str, person: str) -> int:
    return st.session_state.cumulative_tasks.get(f"{task}|{person}", 0)

def inc_cum_task(task: str, person: str):
    k = f"{task}|{person}"
    st.session_state.cumulative_tasks[k] = st.session_state.cumulative_tasks.get(k, 0) + 1

# ─────────────────────────────────────────────────────────────────────────────
#  HELPERS — availability
# ─────────────────────────────────────────────────────────────────────────────
def day_status(person: str, day_idx: int, wk: str) -> str:
    return st.session_state.daily_status.get(f"{wk}|{person}|{day_idx}", "Présent")

def set_day_status(person: str, day_idx: int, wk: str, val: str):
    st.session_state.daily_status[f"{wk}|{person}|{day_idx}"] = val

def is_absent_all_week(person: str, wk: str) -> bool:
    return all(day_status(person, i, wk) != "Présent" for i in range(5))

def violates_no_pair(candidate: str, existing: list) -> bool:
    for a, b in st.session_state.no_pair:
        other = b if candidate == a else (a if candidate == b else None)
        if other and other in existing:
            return True
    return False

def base_pair_partner(person: str):
    for a, b in st.session_state.base_pairs:
        if a == person: return b
        if b == person: return a
    return None

# ─────────────────────────────────────────────────────────────────────────────
#  PLANNING GENERATION
# ─────────────────────────────────────────────────────────────────────────────
def pick_from_pool(pool: list, label: str, counter_fn, already: list) -> str:
    """
    Pick the person with the lowest cumulative counter from pool,
    respecting no_pair and not re-assigning already-assigned people.
    STRICT: only people in pool are considered. Returns VACANT if pool exhausted.
    """
    valid = [
        p for p in pool
        if p not in already
        and not violates_no_pair(p, already)
    ]
    if not valid:
        return "VACANT"
    min_cum = min(counter_fn(label, p) for p in valid)
    candidates = [p for p in valid if counter_fn(label, p) == min_cum]
    return random.choice(candidates)


def generate_planning(wk: str):
    team      = st.session_state.team
    # Available = not absent the whole week
    available = set(p for p in team if not is_absent_all_week(p, wk))

    main_rows = []

    for poste in st.session_state.postes:
        nom  = poste["nom"]
        kind = poste["type"]

        if kind == "fixe":
            person = poste.get("person", "")
            main_rows.append({"Poste": nom, "Assigné": person if person else "VACANT", "Type": "fixe"})

        elif kind == "rotation":
            # STRICT pool: only listed people, filtered by availability
            raw_pool = poste.get("pool", [])
            pool     = [p for p in raw_pool if p in available]
            slots    = max(1, int(poste.get("slots", 1)))

            for slot_i in range(slots):
                already_on = [r["Assigné"] for r in main_rows if r["Poste"] == nom]

                # Base-pair priority
                forced = None
                if slot_i > 0:
                    for ep in already_on:
                        partner = base_pair_partner(ep)
                        if (partner and partner in pool
                                and partner not in already_on
                                and not violates_no_pair(partner, already_on)):
                            forced = partner
                            break

                person = forced if forced else pick_from_pool(pool, nom, get_cum_post, already_on)
                main_rows.append({"Poste": nom, "Assigné": person, "Type": "rotation"})
                if person != "VACANT":
                    inc_cum_post(nom, person)

    # Tasks
    task_rows = []
    for task in st.session_state.tasks:
        raw_pool = st.session_state.task_pools.get(task, [])
        pool     = [p for p in raw_pool if p in available] if raw_pool else list(available)
        person   = pick_from_pool(pool, task, get_cum_task, [])
        task_rows.append({"Tâche": task, "Assigné": person})
        if person != "VACANT":
            inc_cum_task(task, person)

    df_main  = pd.DataFrame(main_rows)  if main_rows  else pd.DataFrame(columns=["Poste","Assigné","Type"])
    df_tasks = pd.DataFrame(task_rows)  if task_rows  else pd.DataFrame(columns=["Tâche","Assigné"])

    st.session_state.week_history[wk] = {
        "main":  df_main.to_dict("records"),
        "tasks": df_tasks.to_dict("records"),
    }
    return df_main, df_tasks

# ─────────────────────────────────────────────────────────────────────────────
#  EXCEL EXPORT
# ─────────────────────────────────────────────────────────────────────────────
def build_excel(df_main, df_tasks, label):
    wb = Workbook(); ws = wb.active; ws.title = "Planning"
    BLUE,DARK,GREY,WHITE = "2563EB","1A1D27","374151","FFFFFF"

    def hdr(cell, bg=BLUE):
        cell.font=Font(bold=True,color=WHITE,size=11)
        cell.fill=PatternFill("solid",fgColor=bg)
        cell.alignment=Alignment(horizontal="center",vertical="center")

    def section(text, ncols=2):
        ws.append([text]+[""]*(ncols-1))
        r=ws.max_row
        ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=ncols)
        c=ws.cell(r,1)
        c.font=Font(bold=True,color=WHITE,size=12)
        c.fill=PatternFill("solid",fgColor=DARK)
        c.alignment=Alignment(horizontal="left",vertical="center")
        ws.row_dimensions[r].height=22

    def table_hdr(cols):
        ws.append(cols)
        for ci in range(1,len(cols)+1): hdr(ws.cell(ws.max_row,ci),GREY)

    ws.merge_cells("A1:C1"); ws["A1"]=f"Planning semaine du {label}"
    hdr(ws["A1"]); ws.row_dimensions[1].height=28
    ws.append(["Généré le:",date.today().strftime("%d/%m/%Y")]); ws.append([])

    section("📌 Postes principaux",2); table_hdr(["Poste","Assigné"])
    d=df_main[["Poste","Assigné"]] if not df_main.empty else pd.DataFrame(columns=["Poste","Assigné"])
    for _,row in d.iterrows(): ws.append([row["Poste"],row["Assigné"]])
    ws.append([])

    section("📋 Tâches annexes",2); table_hdr(["Tâche","Assigné"])
    if not df_tasks.empty:
        for _,row in df_tasks.iterrows(): ws.append([row["Tâche"],row["Assigné"]])
    ws.append([])

    section("🗂️ Configuration des postes",3); table_hdr(["Poste","Type","Détail"])
    for p in st.session_state.postes:
        detail = p.get("person","") if p["type"]=="fixe" else ", ".join(p.get("pool",[]))
        ws.append([p["nom"], p["type"], detail])
    ws.append([])

    section("🚀 Opérations ponctuelles",2); table_hdr(["Opération","Personnes"])
    for op,persons in st.session_state.operations.items(): ws.append([op,", ".join(persons)])
    ws.append([])

    section("✅ Binômes de base",2); table_hdr(["Personne A","Personne B"])
    for a,b in st.session_state.base_pairs: ws.append([a,b])
    ws.append([])

    section("🚫 Binômes interdits",2); table_hdr(["Personne A","Personne B"])
    for a,b in st.session_state.no_pair: ws.append([a,b])

    for col in ws.columns:
        real=[c for c in col if not isinstance(c,MergedCell)]
        if not real: continue
        mx=max((len(str(c.value or "")) for c in real),default=10)
        ws.column_dimensions[real[0].column_letter].width=min(mx+4,50)

    buf=BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()

# ─────────────────────────────────────────────────────────────────────────────
#  SIDEBAR — CONFIGURATION
# ─────────────────────────────────────────────────────────────────────────────
def reset_planning():
    st.session_state.df_main       = pd.DataFrame()
    st.session_state.df_tasks      = pd.DataFrame()
    st.session_state.last_week_key = ""
    st.session_state.editor_version += 1

with st.sidebar:
    st.markdown("## ⚙️ Paramétrage")

    # ── Équipe ────────────────────────────────────────────────────────────────
    with st.expander("👥 Équipe", expanded=True):
        team_raw = st.text_input(
            "Membres (séparés par virgule)",
            value=", ".join(st.session_state.team),
            placeholder="Alice, Bob, Charlie…"
        )

    # ── Postes ────────────────────────────────────────────────────────────────
    with st.expander("📋 Postes", expanded=True):
        st.markdown('<div class="tip-box">💡 <b>Fixe</b> = toujours la même personne.<br>🔄 <b>Rotation</b> = pool strict de personnes habilitées, rotation équitable. Personne hors pool ne peut pas apparaître.</div>', unsafe_allow_html=True)

        postes_edit = []
        for i, p in enumerate(st.session_state.postes):
            st.markdown(f"**Poste {i+1}**")
            c1, c2 = st.columns([3,1])
            nom  = c1.text_input("Nom du poste", value=p.get("nom",""), key=f"pnom_{i}", placeholder="Ex: Kerakoll")
            kind = c2.selectbox("Type", ["fixe","rotation"],
                                index=0 if p.get("type","fixe")=="fixe" else 1,
                                key=f"pkind_{i}", label_visibility="visible")

            if kind == "fixe":
                person = st.selectbox(
                    "Personne dédiée",
                    options=[""] + st.session_state.team,
                    index=([""] + st.session_state.team).index(p.get("person",""))
                          if p.get("person","") in st.session_state.team else 0,
                    key=f"pperson_{i}"
                )
                postes_edit.append({"nom": nom, "type": "fixe", "slots": 1,
                                    "person": person, "pool": []})
            else:
                slots = st.number_input("Nb de personnes à assigner", min_value=1, max_value=20,
                                        value=int(p.get("slots",1)), key=f"pslots_{i}")
                pool_str = st.text_area(
                    "Pool (une personne par ligne ou séparées par virgule)",
                    value="\n".join(p.get("pool",[])),
                    height=90, key=f"ppool_{i}",
                    placeholder="Alice\nBob\nCharlie"
                )
                pool = parse_list(pool_str)
                # Warn if pool members not in team
                unknown = [x for x in pool if x and x not in st.session_state.team]
                if unknown:
                    st.warning(f"⚠️ Hors équipe : {', '.join(unknown)}")
                postes_edit.append({"nom": nom, "type": "rotation", "slots": slots,
                                    "person": "", "pool": pool})

            col_del = st.columns([4,1])[1]
            if col_del.button("🗑️", key=f"pdel_{i}", help="Supprimer ce poste"):
                st.session_state.postes.pop(i)
                reset_planning(); save_data(); st.rerun()

            st.markdown("---")

        if st.button("➕ Ajouter un poste", use_container_width=True):
            st.session_state.postes.append({"nom":"","type":"rotation","slots":1,"person":"","pool":[]})
            save_data(); st.rerun()

    # ── Tâches ────────────────────────────────────────────────────────────────
    with st.expander("📝 Tâches annexes"):
        tasks_input = st.text_area(
            "Une tâche par ligne",
            value="\n".join(st.session_state.tasks),
            height=80, placeholder="Nettoyage\nInventaire"
        )
        task_pools_input = st.text_area(
            "Pool par tâche (tâche:Nom1,Nom2 — vide = toute l'équipe)",
            value="\n".join(f"{k}:{','.join(v)}" for k,v in st.session_state.task_pools.items()),
            height=80, placeholder="Nettoyage:Alice,Bob"
        )

    # ── Binômes ───────────────────────────────────────────────────────────────
    with st.expander("🤝 Binômes"):
        base_pairs_input = st.text_area(
            "✅ Binômes de base (Nom1,Nom2 — toujours ensemble en priorité)",
            value="\n".join(",".join(p) for p in st.session_state.base_pairs),
            height=70, placeholder="Alice,Bob"
        )
        no_pair_input = st.text_area(
            "🚫 Binômes interdits (Nom1,Nom2)",
            value="\n".join(",".join(p) for p in st.session_state.no_pair),
            height=70, placeholder="Charlie,David"
        )

    # ── Opérations ────────────────────────────────────────────────────────────
    with st.expander("🚀 Opérations ponctuelles"):
        operations_input = st.text_area(
            "Format : Opération:Nom1,Nom2",
            value="\n".join(f"{k}:{','.join(v)}" for k,v in st.session_state.operations.items()),
            height=80, placeholder="Container 40:Alice,Bob"
        )

    st.markdown("")
    if st.button("✅ Valider la configuration", use_container_width=True):
        st.session_state.team       = [x.strip() for x in team_raw.split(",") if x.strip()]
        st.session_state.postes     = postes_edit
        st.session_state.tasks      = [x.strip() for x in tasks_input.splitlines() if x.strip()]
        st.session_state.task_pools = parse_dict_list(task_pools_input)
        st.session_state.operations = parse_dict_list(operations_input)
        st.session_state.base_pairs = parse_pairs(base_pairs_input)
        st.session_state.no_pair    = parse_pairs(no_pair_input)
        reset_planning()
        save_data()
        st.toast("Configuration mise à jour ✅")
        st.rerun()

    st.divider()
    if st.button("🗑️ Réinitialiser les compteurs", use_container_width=True):
        st.session_state.cumulative_posts = {}
        st.session_state.cumulative_tasks = {}
        st.session_state.week_history     = {}
        save_data(); st.toast("Compteurs réinitialisés ✅")

# ─────────────────────────────────────────────────────────────────────────────
#  CURRENT WEEK
# ─────────────────────────────────────────────────────────────────────────────
monday     = current_monday()
friday     = monday + timedelta(days=4)
week_dates = [monday + timedelta(days=i) for i in range(5)]
wk         = week_key(monday)
week_label = f"{monday.strftime('%d/%m/%Y')} → {friday.strftime('%d/%m/%Y')}"

# ─────────────────────────────────────────────────────────────────────────────
#  HEADER + NAVIGATION
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("# 📦 Planning Logistique")

c_prev, c_mid, c_next = st.columns([1,4,1])
with c_prev:
    if st.button("⬅️ Semaine préc.", use_container_width=True):
        st.session_state.week_offset -= 1
        reset_planning(); save_data(); st.rerun()
with c_next:
    if st.button("Semaine suiv. ➡️", use_container_width=True):
        st.session_state.week_offset += 1
        reset_planning(); save_data(); st.rerun()

c_mid.markdown(f"""
<div style="background:#1a1d27;border:1px solid #2a2d3e;border-radius:12px;padding:14px 24px;margin:4px 0;">
  <span style="color:#4b5563;font-size:.68rem;text-transform:uppercase;letter-spacing:.1em;">Semaine en cours</span><br>
  <span style="font-size:1.25rem;font-weight:800;">🗓️ {week_label}</span>
  <span style="margin-left:14px;font-family:'DM Mono',monospace;font-size:.72rem;
    color:#2563eb;background:#0f1f3d;padding:3px 10px;border-radius:20px;">{wk}</span>
</div>""", unsafe_allow_html=True)

absent_all  = {p for p in st.session_state.team if is_absent_all_week(p,wk)}
present_cnt = len(st.session_state.team) - len(absent_all)
n_posts     = sum(1 if p["type"]=="fixe" else p.get("slots",1) for p in st.session_state.postes)

m1,m2,m3,m4 = st.columns(4)
for col,val,lbl,color in [
    (m1,len(st.session_state.team),"Membres équipe","#2563eb"),
    (m2,present_cnt,               "Présents",      "#4ade80"),
    (m3,len(absent_all),           "Absences",      "#f87171"),
    (m4,n_posts,                   "Slots postes",  "#fbbf24"),
]:
    col.markdown(f'<div class="metric-card"><div class="val" style="color:{color}">{val}</div>'
                 f'<div class="lbl">{lbl}</div></div>', unsafe_allow_html=True)

st.markdown("<br>", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
#  RESTORE PLANNING FROM HISTORY (no auto-generate)
#  The planning is ONLY generated when the user clicks "Régénérer".
#  On page load / week change, we restore from saved history so everyone
#  sees the exact same planning.
# ─────────────────────────────────────────────────────────────────────────────
if st.session_state.last_week_key != wk:
    # Week changed — try to restore saved planning for this week
    st.session_state.df_main  = pd.DataFrame()
    st.session_state.df_tasks = pd.DataFrame()
    st.session_state.editor_version += 1

    if wk in st.session_state.week_history:
        # Restore exactly what was saved — same for everyone
        saved = st.session_state.week_history[wk]
        main_records  = saved.get("main", [])
        tasks_records = saved.get("tasks", [])
        if main_records:
            st.session_state.df_main = pd.DataFrame(main_records)
        if tasks_records:
            st.session_state.df_tasks = pd.DataFrame(tasks_records)

    st.session_state.last_week_key = wk

# ─────────────────────────────────────────────────────────────────────────────
#  TABS
# ─────────────────────────────────────────────────────────────────────────────
tab_pres,tab_plan,tab_cal,tab_chg,tab_ops,tab_hist,tab_backup = st.tabs([
    "👥 Présence","📋 Planning","📅 Calendrier",
    "⚖️ Charges","🚀 Opérations","📊 Historique","☁️ Sauvegarde"
])

# ══════════════════════════════════════════════════════════════════════════════
#  TAB PRÉSENCE
# ══════════════════════════════════════════════════════════════════════════════
with tab_pres:
    st.markdown("### Gestion des présences")
    st.markdown('<div class="info-box">Statut jour par jour. Sans aucun jour "Présent", la personne est exclue du planning.<br>🟡 <b>Jour Férié</b> : boutons rapides pour appliquer à toute l\'équipe.</div>', unsafe_allow_html=True)

    if not st.session_state.team:
        st.warning("Aucun membre configuré.")
    else:
        # Boutons rapides Jour Férié
        st.markdown("##### Application rapide — Jour Férié")
        fc = st.columns(5)
        for i,(d,wd) in enumerate(zip(DAYS,week_dates)):
            with fc[i]:
                if st.button(f"🟡 {d}\n{wd.strftime('%d/%m')}", key=f"ferie_{i}", use_container_width=True):
                    for p in st.session_state.team: set_day_status(p,i,wk,"Jour Férié")
                    save_data(); st.toast(f"Jour Férié — {d} ✅"); st.rerun()

        st.markdown("---")
        hcols = st.columns([2]+[1]*5)
        hcols[0].markdown('<div class="sec">Personne</div>', unsafe_allow_html=True)
        for i,(d,wd) in enumerate(zip(DAYS,week_dates)):
            hcols[i+1].markdown(
                f'<div class="sec" style="text-align:center">{d}<br>'
                f'<span style="font-family:\'DM Mono\',monospace;font-size:.62rem;color:#6b7280">'
                f'{wd.strftime("%d/%m")}</span></div>', unsafe_allow_html=True)

        for p in st.session_state.team:
            cols = st.columns([2]+[1]*5)
            cols[0].markdown(f"**{p}**")
            for i,d in enumerate(DAYS):
                cur = day_status(p,i,wk)
                with cols[i+1]:
                    s = st.selectbox(f"{p}_{d}_{wk}", STATUTS,
                                     index=STATUTS.index(cur) if cur in STATUTS else 0,
                                     key=f"sel_{wk}_{p}_{i}", label_visibility="collapsed")
                    set_day_status(p,i,wk,s)

        if st.button("💾 Enregistrer les présences", key="save_pres"):
            save_data(); st.toast("Présences enregistrées ✅")

        st.markdown("#### Vue synthétique")
        smap = {"Présent":0,"ABS":1,"Congé":2,"Jour Férié":3}
        z,texts=[],[]
        for p in st.session_state.team:
            rz,rt=[],[]
            for i in range(5):
                s=day_status(p,i,wk); rz.append(smap.get(s,0)); rt.append(s)
            z.append(rz); texts.append(rt)

        fig_p=go.Figure(go.Heatmap(
            z=z, x=[f"{d} {wd.strftime('%d/%m')}" for d,wd in zip(DAYS,week_dates)],
            y=st.session_state.team,
            colorscale=[[0,"#14532d"],[0.33,"#7f1d1d"],[0.66,"#1e1b4b"],[1,"#78350f"]],
            showscale=False, text=texts, texttemplate="%{text}",
            hovertemplate="<b>%{y}</b> — %{x}<br>%{text}<extra></extra>"
        ))
        fig_p.update_layout(paper_bgcolor="rgba(0,0,0,0)",plot_bgcolor="rgba(0,0,0,0)",
            font=dict(color="#e8e8e8",family="Syne"),margin=dict(l=10,r=10,t=10,b=10),
            height=max(200,44*len(st.session_state.team)))
        st.plotly_chart(fig_p, use_container_width=True)

# ══════════════════════════════════════════════════════════════════════════════
#  TAB PLANNING
# ══════════════════════════════════════════════════════════════════════════════
with tab_plan:
    st.markdown("### Planning de la semaine")

    cg,cs,ce = st.columns([2,1,1])
    with cg:
        if st.button("🔁 Régénérer le planning", use_container_width=True):
            st.session_state.df_main,st.session_state.df_tasks = generate_planning(wk)
            st.session_state.last_week_key = wk
            st.session_state.editor_version += 1
            save_data(); st.toast("Planning régénéré ✅"); st.rerun()
    with cs:
        if st.button("💾 Sauvegarder", use_container_width=True):
            st.session_state["pending_save"] = True
            st.rerun()
    with ce:
        if not st.session_state.df_main.empty:
            st.download_button(
                label="📤 Export Excel",
                data=build_excel(st.session_state.df_main,st.session_state.df_tasks,week_label),
                file_name=f"planning_{wk}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

    # ── Conflict detection ────────────────────────────────────────────────────
    conflicts = []
    if not st.session_state.df_main.empty:
        # Build pool map for quick lookup
        pool_map = {}
        for p in st.session_state.postes:
            if p["type"] == "rotation":
                pool_map[p["nom"]] = p.get("pool", [])

        for _,row in st.session_state.df_main.iterrows():
            person,post = row["Assigné"],row["Poste"]
            if person == "VACANT":
                conflicts.append(f"⚠️ **{post}** — aucun membre disponible dans le pool (VACANT).")
                continue
            if is_absent_all_week(person,wk):
                conflicts.append(f"⚠️ **{person}** est absent toute la semaine mais assigné à **{post}**.")
            if post in pool_map and person not in pool_map[post]:
                conflicts.append(f"🚫 **{person}** n'est pas dans le pool de **{post}** (modification manuelle détectée).")

        for p in st.session_state.postes:
            if p["type"] != "rotation": continue
            nom = p["nom"]
            on_post = list(st.session_state.df_main[st.session_state.df_main["Poste"]==nom]["Assigné"])
            for a,b in st.session_state.no_pair:
                if a in on_post and b in on_post:
                    conflicts.append(f"🚫 Binôme interdit **{a}** + **{b}** sur **{nom}**.")

    for c in conflicts:
        st.markdown(f'<div class="conflict-box">{c}</div>', unsafe_allow_html=True)

    # ── Editors ───────────────────────────────────────────────────────────────
    if not st.session_state.df_main.empty:
        all_names = st.session_state.team + ["VACANT"]
        ev = st.session_state.editor_version

        st.markdown("#### Postes principaux")
        edited_main = st.data_editor(
            st.session_state.df_main[["Poste","Assigné"]].copy(),
            use_container_width=True, hide_index=True,
            height=min(600, 55+36*len(st.session_state.df_main)),
            column_config={
                "Poste":   st.column_config.TextColumn("Poste", width="medium", disabled=True),
                "Assigné": st.column_config.SelectboxColumn("Assigné", options=all_names,
                                                             width="medium", required=True),
            },
            key=f"editor_main_{ev}"
        )
        if "Type" in st.session_state.df_main.columns:
            edited_main = edited_main.copy()
            edited_main["Type"] = st.session_state.df_main["Type"].values
        st.session_state.df_main = edited_main

        st.markdown("#### Tâches annexes")
        if not st.session_state.df_tasks.empty:
            edited_tasks = st.data_editor(
                st.session_state.df_tasks.copy(),
                use_container_width=True, hide_index=True,
                height=min(400, 55+36*len(st.session_state.df_tasks)),
                column_config={
                    "Tâche":   st.column_config.TextColumn("Tâche", width="medium", disabled=True),
                    "Assigné": st.column_config.SelectboxColumn("Assigné", options=all_names,
                                                                 width="medium", required=True),
                },
                key=f"editor_tasks_{ev}"
            )
            st.session_state.df_tasks = edited_tasks
        else:
            st.info("Aucune tâche annexe configurée.")

    # ── Flush pending save AFTER editors have updated session_state ─────────────
    if st.session_state.get("pending_save"):
        st.session_state["pending_save"] = False
        save_data()
        st.toast("✅ Planning sauvegardé et envoyé sur le Gist !")

    # ── Résumé des postes ─────────────────────────────────────────────────────
    if st.session_state.postes:
        st.markdown("#### 🗂️ Résumé de la configuration des postes")
        for p in st.session_state.postes:
            nom,kind = p["nom"],p["type"]
            if kind == "fixe":
                person = p.get("person","—")
                st.markdown(
                    f"<div class='post-card post-card-fix'>"
                    f"<span class='badge badge-fix'>FIXE</span> <b>{nom}</b>"
                    f" → <span class='badge badge-ok'>{person}</span></div>",
                    unsafe_allow_html=True)
            else:
                pool    = p.get("pool",[])
                slots   = p.get("slots",1)
                avail   = [x for x in pool if not is_absent_all_week(x,wk)]
                absent  = [x for x in pool if is_absent_all_week(x,wk)]
                badges_parts = []
                for x in pool:
                    cls = "badge-error" if x in absent else "badge-ok"
                    badges_parts.append(f"<span class='badge {cls}'>{x}</span>")
                badges = " ".join(badges_parts)
                empty_msg = "<span style='color:#6b7280'>Pool vide</span>"
                badges_html = badges if badges else empty_msg
                pool_info = f"({len(avail)}/{len(pool)} disponibles)"
                st.markdown(
                    f"<div class='post-card post-card-rot'>"
                    f"<span class='badge badge-rot'>ROTATION x{slots}</span> <b>{nom}</b>"
                    f" <span style='color:#4b5563;font-size:.72rem'>{pool_info}</span>"
                    f"<br><div style='margin-top:6px'>{badges_html}</div>"
                    f"</div>",
                    unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
#  TAB CALENDRIER
# ══════════════════════════════════════════════════════════════════════════════
with tab_cal:
    st.markdown("### Vue Calendrier")

    STATUS_STYLE = {
        "ABS":        ("#2d1515","#f87171","❌"),
        "Congé":      ("#1e1b4b","#a5b4fc","🏖️"),
        "Jour Férié": ("#2d1f00","#fcd34d","🟡"),
    }

    if st.session_state.df_main.empty:
        st.info("Générez d'abord le planning dans l'onglet Planning.")
    else:
        assign_map: dict = defaultdict(list)
        for _,row in st.session_state.df_main.iterrows():
            if row["Assigné"] != "VACANT":
                assign_map[row["Assigné"]].append(row["Poste"])
        if not st.session_state.df_tasks.empty:
            for _,row in st.session_state.df_tasks.iterrows():
                if row["Assigné"] != "VACANT":
                    assign_map[row["Assigné"]].append(f"[{row['Tâche']}]")

        all_post_names = [p["nom"] for p in st.session_state.postes] + st.session_state.tasks
        color_map = {n: POST_COLORS[i%len(POST_COLORS)] for i,n in enumerate(all_post_names)}

        hcols = st.columns([2]+[1]*5)
        hcols[0].markdown('<div class="sec">Personne</div>', unsafe_allow_html=True)
        for i,(d,wd) in enumerate(zip(DAYS,week_dates)):
            hcols[i+1].markdown(
                f'<div class="sec" style="text-align:center">{d}<br>'
                f'<span style="font-family:\'DM Mono\',monospace;font-size:.62rem;color:#6b7280">'
                f'{wd.strftime("%d/%m")}</span></div>', unsafe_allow_html=True)

        for p in st.session_state.team:
            cols = st.columns([2]+[1]*5)
            assignments = assign_map.get(p,[])
            charge_n = len(assignments)
            bc = "badge-ok" if charge_n<=2 else ("badge-warn" if charge_n<=4 else "badge-error")
            cols[0].markdown(
                f"<div style='padding:8px 0'><b>{p}</b> "
                f"<span class='badge {bc}'>{charge_n} poste{'s' if charge_n!=1 else ''}</span></div>",
                unsafe_allow_html=True)

            for i in range(5):
                status = day_status(p,i,wk)
                with cols[i+1]:
                    if status in STATUS_STYLE:
                        bg,fc,emo = STATUS_STYLE[status]
                        st.markdown(
                            f'<div style="background:{bg};border-radius:8px;padding:7px;'
                            f'text-align:center;font-size:.72rem;font-weight:700;color:{fc};margin:2px 0;">'
                            f'{emo} {status}</div>', unsafe_allow_html=True)
                    else:
                        if assignments:
                            html=""
                            for a in assignments:
                                clean=a.strip("[]"); c=color_map.get(clean,"#2563eb")
                                is_t=a.startswith("[")
                                html+=(f'<div style="background:{c}22;border:1px solid {c}55;'
                                       f'border-radius:6px;padding:3px 6px;text-align:center;'
                                       f'font-size:.65rem;font-weight:700;color:{c};margin:2px 0;'
                                       f'font-family:\'DM Mono\',monospace;">'
                                       f'{"📋" if is_t else "📍"} {clean}</div>')
                            st.markdown(html, unsafe_allow_html=True)
                        else:
                            st.markdown(
                                '<div style="background:#1a1d27;border:1px dashed #374151;'
                                'border-radius:8px;padding:8px;text-align:center;'
                                'font-size:.7rem;color:#4b5563;margin:2px 0;">—</div>',
                                unsafe_allow_html=True)

        st.markdown("#### Heatmap de charge")
        hz,ht=[],[]
        snum={"Présent":0,"ABS":-1,"Congé":-2,"Jour Férié":-3}
        for p in st.session_state.team:
            rz,rt=[],[]
            n=len(assign_map.get(p,[]))
            for i in range(5):
                s=day_status(p,i,wk)
                if s!="Présent": rz.append(snum.get(s,-1)); rt.append(s)
                else: rz.append(n); rt.append(f"{n} poste(s)")
            hz.append(rz); ht.append(rt)

        fig_h=go.Figure(go.Heatmap(
            z=hz, x=[f"{d} {wd.strftime('%d/%m')}" for d,wd in zip(DAYS,week_dates)],
            y=st.session_state.team,
            colorscale=[[0,"#1e1b4b"],[0.2,"#78350f"],[0.35,"#2d1515"],
                        [0.5,"#14532d"],[0.65,"#4ade80"],[0.82,"#fbbf24"],[1,"#dc2626"]],
            showscale=True, text=ht, texttemplate="%{text}",
            hovertemplate="<b>%{y}</b> — %{x}<br>%{text}<extra></extra>"
        ))
        fig_h.update_layout(paper_bgcolor="rgba(0,0,0,0)",plot_bgcolor="rgba(0,0,0,0)",
            font=dict(color="#e8e8e8",family="Syne"),margin=dict(l=10,r=10,t=10,b=10),
            height=max(250,52*len(st.session_state.team)))
        st.plotly_chart(fig_h, use_container_width=True)

# ══════════════════════════════════════════════════════════════════════════════
#  TAB CHARGES
# ══════════════════════════════════════════════════════════════════════════════
with tab_chg:
    st.markdown("### Répartition des charges")

    if st.session_state.team:
        charge={p:0 for p in st.session_state.team}
        if not st.session_state.df_main.empty:
            for p in st.session_state.df_main["Assigné"]:
                if p in charge: charge[p]+=1
        if not st.session_state.df_tasks.empty:
            for p in st.session_state.df_tasks["Assigné"]:
                if p in charge: charge[p]+=1

        charge_df=pd.DataFrame({"Opérateur":list(charge.keys()),"Charge":list(charge.values())})
        max_c=max(charge.values(),default=1)
        filtre=st.slider("Charge minimale à afficher",0,max(max_c,1),0)
        filtered=charge_df[charge_df["Charge"]>=filtre]

        colors=["#4ade80" if v<=2 else "#fbbf24" if v<=4 else "#f87171" for v in filtered["Charge"]]
        fig_b=go.Figure(go.Bar(x=filtered["Opérateur"],y=filtered["Charge"],
            marker_color=colors,text=filtered["Charge"],textposition="outside"))
        fig_b.update_layout(paper_bgcolor="rgba(0,0,0,0)",plot_bgcolor="rgba(0,0,0,0)",
            font=dict(color="#e8e8e8",family="Syne"),margin=dict(l=10,r=10,t=10,b=10),
            yaxis=dict(gridcolor="#1f2937"),height=340)
        st.plotly_chart(fig_b,use_container_width=True)

        st.markdown("#### Compteurs de rotation cumulatifs")
        st.markdown('<div class="info-box">Compteur bas = prioritaire à la prochaine rotation dans ce pool.</div>', unsafe_allow_html=True)
        cum_rows=[]
        rot_postes=[p for p in st.session_state.postes if p["type"]=="rotation"]
        for person in st.session_state.team:
            row={"Personne":person}
            for p in rot_postes:
                pool=p.get("pool",[])
                row[f"📍 {p['nom']}"]=get_cum_post(p["nom"],person) if (not pool or person in pool) else "—"
            for task in st.session_state.tasks:
                tp=st.session_state.task_pools.get(task,[])
                row[f"📋 {task}"]=get_cum_task(task,person) if (not tp or person in tp) else "—"
            cum_rows.append(row)
        if cum_rows:
            st.dataframe(pd.DataFrame(cum_rows),use_container_width=True,hide_index=True)
    else:
        st.info("Configurez votre équipe dans le panneau latéral.")

# ══════════════════════════════════════════════════════════════════════════════
#  TAB OPÉRATIONS
# ══════════════════════════════════════════════════════════════════════════════
with tab_ops:
    st.markdown("### Opérations ponctuelles")
    st.markdown('<div class="info-box">Ces opérations ne font pas partie de la rotation automatique.</div>', unsafe_allow_html=True)

    if st.session_state.operations:
        ops_rows=[{"Opération":op,"Personnes":", ".join(p),"Nb":len(p)}
                  for op,p in st.session_state.operations.items()]
        st.dataframe(pd.DataFrame(ops_rows),use_container_width=True,hide_index=True)
    else:
        st.info("Aucune opération configurée.")

    col_b1,col_b2=st.columns(2)
    with col_b1:
        st.markdown("**✅ Binômes de base**")
        if st.session_state.base_pairs:
            for a,b in st.session_state.base_pairs:
                st.markdown(f"<span class='badge badge-ok'>{a}</span> + <span class='badge badge-ok'>{b}</span>",unsafe_allow_html=True)
        else: st.caption("Aucun.")
    with col_b2:
        st.markdown("**🚫 Binômes interdits**")
        if st.session_state.no_pair:
            for a,b in st.session_state.no_pair:
                st.markdown(f"<span class='badge badge-error'>{a}</span> + <span class='badge badge-error'>{b}</span>",unsafe_allow_html=True)
        else: st.caption("Aucun.")

# ══════════════════════════════════════════════════════════════════════════════
#  TAB HISTORIQUE
# ══════════════════════════════════════════════════════════════════════════════
with tab_hist:
    st.markdown("### Historique des plannings")

    if not st.session_state.week_history:
        st.info("Aucun historique disponible.")
    else:
        weeks_sorted=sorted(st.session_state.week_history.keys(),reverse=True)
        sel=st.selectbox("Sélectionner une semaine",weeks_sorted)
        if sel:
            hist=st.session_state.week_history[sel]
            ch1,ch2=st.columns(2)
            with ch1:
                st.markdown("**Postes principaux**")
                dm=pd.DataFrame(hist.get("main",[]))
                if not dm.empty:
                    st.dataframe(dm[[c for c in ["Poste","Assigné"] if c in dm.columns]],
                                 use_container_width=True,hide_index=True)
                else: st.caption("Vide")
            with ch2:
                st.markdown("**Tâches annexes**")
                dt=pd.DataFrame(hist.get("tasks",[]))
                if not dt.empty: st.dataframe(dt,use_container_width=True,hide_index=True)
                else: st.caption("Vide")

        if len(st.session_state.week_history)>1:
            st.markdown("#### Fréquence d'attribution par poste")
            freq: dict=defaultdict(lambda:defaultdict(int))
            for wkd in st.session_state.week_history.values():
                for r in wkd.get("main",[]):
                    if r.get("Assigné","VACANT")!="VACANT":
                        freq[r["Poste"]][r["Assigné"]]+=1
            if freq:
                post_sel=st.selectbox("Poste à analyser",list(freq.keys()),key="hist_post")
                pers=list(freq[post_sel].keys())
                cnts=[freq[post_sel][p] for p in pers]
                fig_f=go.Figure(go.Bar(x=pers,y=cnts,marker_color="#7c3aed",text=cnts,textposition="outside"))
                fig_f.update_layout(paper_bgcolor="rgba(0,0,0,0)",plot_bgcolor="rgba(0,0,0,0)",
                    font=dict(color="#e8e8e8",family="Syne"),
                    title=dict(text=f"Rotations — {post_sel}",font=dict(color="#e8e8e8")),
                    yaxis=dict(gridcolor="#1f2937",title="Fois assigné"),
                    margin=dict(l=10,r=10,t=40,b=10),height=320)
                st.plotly_chart(fig_f,use_container_width=True)


# ══════════════════════════════════════════════════════════════════════════════
#  TAB SAUVEGARDE
# ══════════════════════════════════════════════════════════════════════════════
with tab_backup:
    st.markdown("### ☁️ Sauvegarde & Restauration")

    # ── Section 1 : Export / Import JSON manuel ───────────────────────────────
    st.markdown("#### 💾 Export / Import manuel")
    st.markdown('''<div class="info-box">
    L'export JSON contient <b>toute</b> votre configuration : équipe, postes, pools,
    présences, compteurs de rotation, historique. Téléchargez-le régulièrement comme
    sauvegarde, et réimportez-le après un redémarrage du serveur pour retrouver
    exactement l'état précédent.
    </div>''', unsafe_allow_html=True)

    col_exp, col_imp = st.columns(2)

    with col_exp:
        st.markdown("##### 📤 Exporter la configuration")
        export_data = build_save_dict()
        export_bytes = json.dumps(export_data, indent=2, ensure_ascii=False).encode("utf-8")
        st.download_button(
            label="⬇️ Télécharger planning_backup.json",
            data=export_bytes,
            file_name=f"planning_backup_{date.today().strftime('%Y%m%d')}.json",
            mime="application/json",
            use_container_width=True
        )
        st.caption("Sauvegardez ce fichier sur votre PC ou dans un Drive partagé.")

    with col_imp:
        st.markdown("##### 📥 Importer une configuration")
        uploaded = st.file_uploader(
            "Choisir un fichier planning_backup.json",
            type=["json"],
            key="json_uploader"
        )
        if uploaded is not None:
            try:
                imported = json.load(uploaded)
                # Preview
                team_preview = imported.get("team", [])
                postes_preview = imported.get("postes", [])
                st.markdown(
                    f"<div class='tip-box'>✅ Fichier valide détecté<br>"
                    f"👥 Équipe : {len(team_preview)} membres<br>"
                    f"📋 Postes : {len(postes_preview)}<br>"
                    f"📅 Semaines d'historique : {len(imported.get('week_history', {}))}"
                    f"</div>",
                    unsafe_allow_html=True
                )
                if st.button("✅ Charger cette configuration", use_container_width=True, key="do_import"):
                    bm = imported.get("base_monday", str(_monday(date.today())))
                    st.session_state.base_monday      = date.fromisoformat(bm)
                    st.session_state.week_offset      = imported.get("week_offset", 0)
                    st.session_state.team             = imported.get("team", [])
                    st.session_state.postes           = migrate_old_format(imported)
                    st.session_state.tasks            = imported.get("tasks", [])
                    st.session_state.task_pools       = imported.get("task_pools", {})
                    st.session_state.operations       = imported.get("operations", {})
                    st.session_state.daily_status     = imported.get("daily_status", {})
                    st.session_state.no_pair          = [tuple(p) for p in imported.get("no_pair", [])]
                    st.session_state.base_pairs       = [tuple(p) for p in imported.get("base_pairs", [])]
                    st.session_state.cumulative_posts = imported.get("cumulative_posts", {})
                    st.session_state.cumulative_tasks = imported.get("cumulative_tasks", {})
                    st.session_state.week_history     = imported.get("week_history", {})
                    reset_planning()
                    save_data()
                    st.toast("Configuration importée et sauvegardée ✅")
                    st.rerun()
            except Exception as e:
                st.error(f"Fichier invalide : {e}")

    st.divider()

    # ── Section 2 : GitHub Gist (persistance automatique) ────────────────────
    st.markdown("#### ☁️ Persistance automatique via GitHub Gist")
    st.markdown('''<div class="info-box">
    Configurez un <b>GitHub Gist privé</b> pour que les données soient sauvegardées
    automatiquement dans le cloud à chaque modification. Gratuit, sans limite.
    Les données survivent aux redémarrages du serveur Streamlit.
    </div>''', unsafe_allow_html=True)

    token_ok  = bool(_gist_token())
    gist_id_ok = bool(_gist_id())

    if token_ok and gist_id_ok:
        st.markdown('<div class="tip-box">✅ GitHub Gist configuré et actif — sauvegarde automatique en cours.</div>', unsafe_allow_html=True)

        col_push, col_pull = st.columns(2)
        with col_push:
            if st.button("⬆️ Forcer la sauvegarde vers Gist", use_container_width=True):
                ok = _gist_push(build_save_dict())
                st.toast("✅ Sauvegardé sur Gist !" if ok else "❌ Échec — vérifiez le token.")
        with col_pull:
            if st.button("⬇️ Restaurer depuis Gist", use_container_width=True):
                pulled = _gist_pull()
                if pulled:
                    bm = pulled.get("base_monday", str(_monday(date.today())))
                    st.session_state.base_monday      = date.fromisoformat(bm)
                    st.session_state.week_offset      = pulled.get("week_offset", 0)
                    st.session_state.team             = pulled.get("team", [])
                    st.session_state.postes           = migrate_old_format(pulled)
                    st.session_state.tasks            = pulled.get("tasks", [])
                    st.session_state.task_pools       = pulled.get("task_pools", {})
                    st.session_state.operations       = pulled.get("operations", {})
                    st.session_state.daily_status     = pulled.get("daily_status", {})
                    st.session_state.no_pair          = [tuple(p) for p in pulled.get("no_pair", [])]
                    st.session_state.base_pairs       = [tuple(p) for p in pulled.get("base_pairs", [])]
                    st.session_state.cumulative_posts = pulled.get("cumulative_posts", {})
                    st.session_state.cumulative_tasks = pulled.get("cumulative_tasks", {})
                    st.session_state.week_history     = pulled.get("week_history", {})
                    reset_planning()
                    save_data()
                    st.toast("✅ Restauré depuis Gist !")
                    st.rerun()
                else:
                    st.error("Impossible de récupérer les données depuis Gist.")
    else:
        st.markdown('''<div class="tip-box">
        📋 <b>Guide de configuration (5 minutes) :</b><br><br>
        <b>Étape 1</b> — Créer un token GitHub :<br>
        → github.com → Settings → Developer settings → Personal access tokens → Tokens (classic)<br>
        → "Generate new token" → cochez uniquement <b>gist</b> → Generate → Copiez le token<br><br>
        <b>Étape 2</b> — Créer un Gist privé vide :<br>
        → gist.github.com → Description : "Planning Logistique" → Contenu : <code>{}</code><br>
        → Cliquez "Create secret gist" → Copiez l'ID dans l'URL (32 caractères)<br><br>
        <b>Étape 3</b> — Configurer dans Streamlit Cloud :<br>
        → Sur share.streamlit.io → votre app → Settings → Secrets<br>
        → Collez exactement :<br>
        <code>GIST_TOKEN = "ghp_votre_token_ici"<br>
        GIST_ID = "votre_gist_id_ici"</code><br><br>
        <b>Étape 4</b> — Redémarrez l'app → La sauvegarde est automatique !
        </div>''', unsafe_allow_html=True)

    st.divider()

    # ── Section 3 : Statut de l'environnement ────────────────────────────────
    st.markdown("#### 🔍 Statut de l'environnement")
    env_rows = [
        {"Paramètre": "Fichier local (planning_data.json)", "Statut": "✅ Présent" if os.path.exists(SAVE_FILE) else "⚠️ Absent (normal sur Streamlit Cloud)"},
        {"Paramètre": "Token GitHub Gist",  "Statut": "✅ Configuré" if token_ok  else "❌ Non configuré"},
        {"Paramètre": "ID du Gist",         "Statut": "✅ Configuré" if gist_id_ok else "❌ Non configuré"},
        {"Paramètre": "Persistance active", "Statut": "✅ Oui (Gist)" if (token_ok and gist_id_ok) else "⚠️ Non — export manuel recommandé"},
        {"Paramètre": "Membres configurés", "Statut": f"{len(st.session_state.team)} membre(s)"},
        {"Paramètre": "Postes configurés",  "Statut": f"{len(st.session_state.postes)} poste(s)"},
        {"Paramètre": "Semaines d'historique", "Statut": f"{len(st.session_state.week_history)}"},
    ]
    st.dataframe(pd.DataFrame(env_rows), use_container_width=True, hide_index=True)

# ─────────────────────────────────────────────────────────────────────────────
#  FOOTER
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("---")
st.markdown(
    '<div style="text-align:center;color:#374151;font-size:.68rem;'
    'font-family:\'DM Mono\',monospace;letter-spacing:.1em;">'
    'PLANNING LOGISTIQUE · Postes Fixe & Rotation · Pool strict · Rotation équitable · Export Excel'
    '</div>', unsafe_allow_html=True)